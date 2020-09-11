import openpyxl
path= "C:\\python\\data\\data.xlsx"
workbook= openpyxl.load_workbook(path)
sheet=workbook.active   # single sheet
 # sheet=workbook.get_sheet_by_name("Sheet1")  multiple sheet
rows=sheet.max_row
colm=sheet.max_column
print(rows,colm)
for r in range(1,rows+1):
    for c in range(1,colm+1):
        print(sheet.cell(row=r,column=c).value, end="    ")
    print()