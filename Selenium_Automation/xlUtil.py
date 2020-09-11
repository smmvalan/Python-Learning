import openpyxl
def getRowCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)

def getColumCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    return(sheet.max_column)

def readDataFile (file,sheetName,rowNum,columNo,):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rowNum,column=columNo).value

def writeDataFile (file,sheetName,rowNum,columNo,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rowNum,column=columNo).value=data
    workbook.save(file)
